"""Excel export service for intel items."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


SEVERITY_COLORS = {
    "critical": "FF0000",
    "high": "FF6600",
    "medium": "FFCC00",
    "low": "33CC33",
    "info": "3399FF",
    "unknown": "CCCCCC",
}

HEADERS = [
    "Title",
    "Severity",
    "Risk Score",
    "CVE IDs",
    "Source",
    "Feed Type",
    "Asset Type",
    "Published At",
    "Ingested At",
    "First Appeared (Source)",
    "Confidence",
    "TLP",
    "Impacted Assets / Products",
    "Exploitability Score",
    "Exploit Available",
    "KEV Listed",
    "Tags",
    "Geo",
    "Industries",
    "Related IOC Count",
    "Source URL",
    "Summary",
]


def export_to_excel(items: list[dict]) -> io.BytesIO:
    """Generate an Excel file from intel items."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Threat Intel"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row_idx, item in enumerate(items, 2):
        sev = item.get("severity", "unknown")
        sev_fill = PatternFill(
            start_color=SEVERITY_COLORS.get(sev, "CCCCCC"),
            end_color=SEVERITY_COLORS.get(sev, "CCCCCC"),
            fill_type="solid",
        )

        row_data = [
            item.get("title", ""),
            sev.upper(),
            item.get("risk_score", 0),
            ", ".join(item.get("cve_ids", [])),
            item.get("source_name", ""),
            item.get("feed_type", ""),
            item.get("asset_type", ""),
            _fmt_dt(item.get("published_at")),
            _fmt_dt(item.get("ingested_at")),
            _fmt_dt(item.get("published_at")),  # first appeared = published
            item.get("confidence", 0),
            item.get("tlp", ""),
            ", ".join(item.get("affected_products", [])),
            item.get("exploitability_score", ""),
            "Yes" if item.get("exploit_available") else "No",
            "Yes" if item.get("is_kev") else "No",
            ", ".join(item.get("tags", [])),
            ", ".join(item.get("geo", [])),
            ", ".join(item.get("industries", [])),
            item.get("related_ioc_count", 0),
            item.get("source_url", ""),
            item.get("summary", ""),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col == 2:  # Severity column
                cell.fill = sev_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = min(len(str(cell.value)), 50)
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _fmt_dt(val) -> str:
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S UTC")
    if isinstance(val, str):
        return val
    return str(val)
